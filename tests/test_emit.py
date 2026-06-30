import json
import os

from drhp_pipeline.contract import (
    Dashboard, Filing, Financials, Meta, MetricValue, Summary,
)
from drhp_pipeline.emit import write_excel_appendix, write_latest


def sample_dashboard():
    f = Filing(
        id="abc", company_name="Acme Ltd", company_name_normalized="acme",
        filing_date="2026-06-30", filing_type="DRHP", stage="DRHP", sector="Materials",
        financials=Financials(
            revenue_fy25=MetricValue(value=800.0, source="DRHP_PDF", confidence="low"),
        ),
        stamps=["FILED_THIS_WEEK"],
    )
    return Dashboard(
        meta=Meta(run_date="2026-06-30", week_start="2026-06-24", week_end="2026-06-30",
                  data_as_of="2026-06-30", snapshot_id="2026-06-30"),
        summary=Summary(new_drhp_count=1), filings=[f],
    )


def test_write_latest_is_valid_contract_json(tmp_path):
    path = str(tmp_path / "latest.json")
    write_latest(sample_dashboard(), path)
    obj = json.load(open(path))
    assert obj["filings"][0]["company_name"] == "Acme Ltd"
    assert obj["filings"][0]["competitor_impact"] is None


def test_excel_appendix_written_with_low_confidence_flag(tmp_path):
    path = str(tmp_path / "appendix.xlsx")
    result = write_excel_appendix(sample_dashboard(), path)
    assert result and os.path.exists(path)
    from openpyxl import load_workbook
    ws = load_workbook(path).active
    header = [c.value for c in ws[1]]
    assert "Company" in header and "Bucket" in header and "Verify" in header
    # low-confidence revenue should be flagged in the Verify column
    verify_idx = header.index("Verify")
    assert "revenue_fy25" in (ws[2][verify_idx].value or "")
