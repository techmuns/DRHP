"""
Stage 8 — Emit.

Writes the current dashboard to `public/data/latest.json` (the file the UI reads) and
refreshes an Excel tracker appendix that mirrors it. The Excel column order is fixed
here so it stays stable run-to-run (and matches the columns the manual tracker used).
"""

from __future__ import annotations

import logging
import os
from typing import List, Optional

from .contract import Dashboard, Filing

log = logging.getLogger("drhp.emit")

LATEST_PATH = "public/data/latest.json"
APPENDIX_PATH = "public/data/tracker_appendix.xlsx"

# Fixed appendix column order (label, extractor).
_COLUMNS = [
    ("Company", lambda f: f.company_name),
    ("Filing Date", lambda f: f.filing_date),
    ("Type", lambda f: f.filing_type),
    ("Stage", lambda f: f.stage),
    ("Sector", lambda f: f.sector),
    ("Sub-sector", lambda f: f.sub_sector),
    ("Issue Type", lambda f: f.issue.type),
    ("Fresh (Cr)", lambda f: f.issue.fresh_cr),
    ("OFS (Cr)", lambda f: f.issue.ofs_cr),
    ("Total Issue (Cr)", lambda f: f.issue.total_cr),
    ("Revenue FY25 (Cr)", lambda f: f.financials.revenue_fy25.value),
    ("Rev Growth %", lambda f: f.financials.rev_growth_pct.value),
    ("PAT FY25 (Cr)", lambda f: f.financials.pat_fy25.value),
    ("PAT Growth %", lambda f: f.financials.pat_growth_pct.value),
    ("PAT Margin %", lambda f: f.financials.pat_margin_pct.value),
    ("EBITDA Margin %", lambda f: f.financials.ebitda_margin_pct.value),
    ("ROE %", lambda f: f.financials.roe_pct.value),
    ("ROCE %", lambda f: f.financials.roce_pct.value),
    ("Debt/Equity", lambda f: f.financials.debt_equity.value),
    ("Promoter %", lambda f: f.financials.promoter_hold_pct.value),
    ("Score", lambda f: f.score.total),
    ("Bucket", lambda f: f.score.bucket),
    ("Stamps", lambda f: ", ".join(f.stamps)),
    ("Data Source", lambda f: _provenance(f)),
    ("Verify", lambda f: _low_confidence_fields(f)),
    ("SEBI URL", lambda f: f.sources.sebi_url),
    ("DRHP PDF", lambda f: f.sources.drhp_pdf_url),
]

# Financial fields surfaced in the provenance / verify columns.
_FIN_FIELDS = [
    "revenue_fy25", "revenue_fy24", "pat_fy25", "pat_fy24", "ebitda_fy25",
    "rev_growth_pct", "pat_growth_pct", "pat_margin_pct", "ebitda_margin_pct",
    "roe_pct", "roce_pct", "debt_equity", "promoter_hold_pct",
]


def _provenance(f: Filing) -> str:
    sources = {getattr(f.financials, name).source for name in _FIN_FIELDS}
    sources.discard(None)
    return ", ".join(sorted(sources)) if sources else "—"


def _low_confidence_fields(f: Filing) -> str:
    """List fields the team should double-check (web-sourced or low confidence)."""
    flagged = []
    for name in _FIN_FIELDS:
        mv = getattr(f.financials, name)
        if mv.value is not None and (mv.source == "WEB" or mv.confidence == "low"):
            flagged.append(name)
    return ", ".join(flagged) if flagged else ""


def write_latest(dashboard: Dashboard, path: str = LATEST_PATH) -> str:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(dashboard.to_json())
    log.info("Wrote %s (%d filings)", path, len(dashboard.filings))
    return path


def write_excel_appendix(dashboard: Dashboard, path: str = APPENDIX_PATH) -> Optional[str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        log.warning("openpyxl not available; skipping Excel appendix.")
        return None

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Tracker"

    header_fill = PatternFill("solid", fgColor="013E37")  # deep green per brand
    header_font = Font(bold=True, color="FFFFFF")
    ws.append([label for label, _ in _COLUMNS])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for f in dashboard.filings:
        ws.append([extractor(f) for _, extractor in _COLUMNS])

    # Readable widths + frozen header.
    for col_idx, (label, _) in enumerate(_COLUMNS, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(11, min(40, len(label) + 4))
    ws.freeze_panes = "A2"

    wb.save(path)
    log.info("Wrote Excel appendix %s", path)
    return path
